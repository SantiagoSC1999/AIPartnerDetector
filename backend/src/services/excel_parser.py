"""Excel parsing and validation service."""

import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = ["id", "partner_name", "institution_type", "country_id"]
OPTIONAL_COLUMNS = ["acronym", "web_page"]


class ExcelParsingError(Exception):
    """Exception raised during Excel parsing."""

    pass


def validate_excel_file(file_content: bytes) -> bool:
    """Validate that the file is a valid Excel file."""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        return True
    except Exception as e:
        raise ExcelParsingError(f"Invalid Excel file: {str(e)}")


def parse_excel_file(file_content: bytes) -> tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse Excel file and return list of records and validation errors.

    Returns:
        Tuple of (records list, errors list)
    """
    errors = []
    records = []

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active

        # Get header row
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value.lower().strip())

        # Validate required columns
        missing_required_columns = set(REQUIRED_COLUMNS) - set(headers)
        if missing_required_columns:
            raise ExcelParsingError(
                f"Missing required columns: {', '.join(missing_required_columns)}"
            )

        # Parse data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                # Skip empty rows
                continue

            record = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    record[header] = row[col_idx]

            # Validate that REQUIRED fields are not empty
            has_errors = False
            for required_field in REQUIRED_COLUMNS:
                if not record.get(required_field):
                    errors.append(f"Row {row_idx}: Missing or empty '{required_field}'")
                    has_errors = True

            # Optional fields can be empty, set to empty string if missing
            for optional_field in OPTIONAL_COLUMNS:
                if optional_field not in record:
                    record[optional_field] = ""

            if not has_errors:
                records.append(record)

    except ExcelParsingError as e:
        errors.append(str(e))
    except Exception as e:
        errors.append(f"Error parsing Excel file: {str(e)}")

    return records, errors
